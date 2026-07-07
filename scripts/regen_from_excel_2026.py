"""
Regenerate the NBMS data files from the authoritative 2026 inventory.
Source: "Bridges and Culverts 2026-FINAL 3.xlsx" (DNR Bridge Section).

Patches public/data/{bridges,culverts,analytics,critical_structures}.json with
the REAL BMS condition ratings so the Overview dashboard, condition distribution
and Critical Structures view are accurate. DNR overall-rating scale (0-10):
  0-1 Critical | 2-3 Poor | 4-5 Marginal | 6 Satisfactory | 7-10 Good
"""
import openpyxl, json, os
from collections import Counter

XL = r"D:\OneDrive\Bridge stuff\Bridges and Culverts 2026-FINAL 3.xlsx"
HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(HERE, "public", "data")
wb = openpyxl.load_workbook(XL, read_only=True, data_only=True)


def clean(v):
    s = str(v).strip() if v is not None else ""
    return "" if s in ("-", "None") else s


def num(v):
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


# App condition scale (0-9, higher = better)
CAT2APP = {
    "Critical": (1, "Critical"), "Poor": (3, "Poor"), "Marginal": (4, "Marginal"),
    "Satisfactory": (6, "Satisfactory"), "Good": (7, "Good"),
}


def cat_norm(v):
    s = clean(v)
    for k in ["Critical", "Poor", "Marginal", "Satisfactory", "Good"]:
        if s.lower().startswith(k.lower()[:5]):
            return k
    if s.lower().startswith("under"):
        return "Under Construction"
    return None


# ── Bridge lookup ────────────────────────────────────────────────────────────
ws = wb["BRIDGES 2026"]
BKEYS, bridge_rows = {}, []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not (clean(r[0]) or clean(r[1]) or clean(r[10])):
        continue
    info = {
        "oldNo": clean(r[0]), "newNo": clean(r[1]), "name": clean(r[10]),
        "linkId": clean(r[53]) or clean(r[13]), "linkName": clean(r[55]) or clean(r[12]),
        "station": clean(r[60]), "region": clean(r[61]),
        "length": num(r[44]), "width": num(r[45]), "comment": clean(r[78]),
        "cat": cat_norm(r[77]), "overall10": num(r[76]),
        "comp": {"approaches": num(r[64]), "roadway": num(r[66]), "substructure": num(r[68]),
                 "superstructure": num(r[70]), "waterway": num(r[72])},
    }
    bridge_rows.append(info)
    for key in (info["oldNo"], info["newNo"]):
        if key:
            BKEYS[key.upper()] = info

# ── Culvert lookup ───────────────────────────────────────────────────────────
ws = wb["MAJOR CULVERTS 2026"]
CKEYS, culvert_rows = {}, []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not (clean(r[0]) or clean(r[1])):
        continue
    info = {"oldNo": clean(r[0]), "newNo": clean(r[1]), "region": clean(r[54]),
            "station": clean(r[53]), "road": clean(r[48]) or clean(r[13]),
            "cat": cat_norm(r[64]), "overall10": num(r[63])}
    culvert_rows.append(info)
    for key in (info["oldNo"], info["newNo"]):
        if key:
            CKEYS[key.upper()] = info


def load(name):
    return json.load(open(os.path.join(DATA, name), encoding="utf-8"))


def save(name, obj):
    json.dump(obj, open(os.path.join(DATA, name), "w", encoding="utf-8"), ensure_ascii=False)


# ── Patch bridges.json ───────────────────────────────────────────────────────
bridges = load("bridges.json")
matched = 0
for b in bridges:
    cand = [clean(b.get("bridge_no")), clean(b.get("original_bridge_no")), clean(b.get("new_bridge_no"))]
    info = next((BKEYS[c.upper()] for c in cand if c and c.upper() in BKEYS), None)
    if not info or not info["cat"] or info["cat"] == "Under Construction":
        if info and info["cat"] == "Under Construction":
            b["overall_rating"] = None
            b["ConditionCategory"] = "Under Construction"
        continue
    matched += 1
    appNum, appLabel = CAT2APP[info["cat"]]
    b["overall_rating"] = appNum
    b["OverallConditionRating"] = appNum
    b["OverallCondition"] = appLabel
    b["ConditionCategory"] = info["cat"]
    c = info["comp"]
    for field, key in (("approaches_rating", "approaches"), ("roadway_rating", "roadway"),
                       ("substructure_rating", "substructure"), ("superstructure_rating", "superstructure"),
                       ("waterway_rating", "waterway")):
        if c[key] is not None:
            b[field] = c[key]
    if info["region"]:
        b["region"] = info["region"]
    if info["station"]:
        b["station"] = info["station"]
        b["maintenanc"] = info["station"]
save("bridges.json", bridges)
print(f"bridges patched: {matched}/{len(bridges)}")

# ── Patch culverts.json ──────────────────────────────────────────────────────
culverts = load("culverts.json")
cm = 0
for cv in culverts:
    cand = [clean(cv.get("CulvertNumber")), clean(cv.get("New Culvert Number")), clean(cv.get("OldCulvertNumber"))]
    info = next((CKEYS[c.upper()] for c in cand if c and c.upper() in CKEYS), None)
    if not info or not info["cat"] or info["cat"] == "Under Construction":
        continue
    cm += 1
    appNum, appLabel = CAT2APP[info["cat"]]
    cv["Overall Rating"] = appNum
    cv["OverallConditionRating"] = appNum
    cv["OverallCondition"] = appLabel
    cv["ConditionCategory"] = info["cat"]
    if info["region"]:
        cv["Region"] = info["region"]
        cv["Maintenance_Region"] = info["region"]
    if info["station"]:
        cv["Maintenance_Station"] = info["station"]
save("culverts.json", culverts)
print(f"culverts patched: {cm}/{len(culverts)}")

# ── Regenerate analytics.json ────────────────────────────────────────────────
analytics = load("analytics.json")
cond, breg, creg, cvcond = Counter(), Counter(), Counter(), Counter()
c_sub, c_sup, c_wat = Counter(), Counter(), Counter()


def band(v):
    if v is None:
        return None
    if v <= 2:
        return "Critical"
    if v == 3:
        return "Poor"
    if v <= 5:
        return "Marginal"
    if v == 6:
        return "Satisfactory"
    return "Good"


for info in bridge_rows:
    if info["cat"] and info["cat"] != "Under Construction":
        cond[CAT2APP[info["cat"]][1]] += 1
        if info["region"]:
            breg[info["region"]] += 1
    for tgt, key in ((c_sub, "substructure"), (c_sup, "superstructure"), (c_wat, "waterway")):
        bb = band(info["comp"][key])
        if bb:
            tgt[bb] += 1
for info in culvert_rows:
    if info["cat"] and info["cat"] != "Under Construction":
        cvcond[CAT2APP[info["cat"]][1]] += 1
        if info["region"]:
            creg[info["region"]] += 1
analytics["condition_overall"] = dict(cond)
analytics["bridges_by_region"] = dict(breg)
analytics["culverts_by_region"] = dict(creg)
analytics["culvert_conditions"] = dict(cvcond)
analytics["condition_substructure"] = dict(c_sub)
analytics["condition_superstructure"] = dict(c_sup)
analytics["condition_waterway"] = dict(c_wat)
analytics["critical_counts"] = {
    "bridges_poor_or_worse": cond["Critical"] + cond["Poor"],
    "bridges_critical": cond["Critical"],
    "culverts_poor_or_worse": cvcond["Critical"] + cvcond["Poor"],
}
save("analytics.json", analytics)
print("condition_overall:", dict(cond))
print("bridges_by_region:", dict(breg))

# ── Regenerate critical_structures.json (Critical + Poor bridges) ────────────
sev = {"Critical": 0, "Poor": 1}
crit = []
for info in sorted(bridge_rows, key=lambda x: sev.get(x["cat"], 9)):
    if info["cat"] in ("Critical", "Poor"):
        crit.append({
            "BridgeNumber": info["newNo"] or info["oldNo"],
            "BridgeName": info["name"],
            "LinkID": info["linkId"],
            "LinkName": info["linkName"],
            "MaintenanceStation": info["station"],
            "BridgeLength": str(info["length"]) if info["length"] is not None else "",
            "BridgeWidth": str(info["width"]) if info["width"] is not None else "",
            "OverallRating": info["cat"],
            "Comment": info["comment"],
        })
save("critical_structures.json", crit)
print(f"critical_structures: {len(crit)} (Critical+Poor bridges)")
